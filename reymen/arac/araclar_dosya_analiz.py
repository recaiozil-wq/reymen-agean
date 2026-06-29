# -*- coding: utf-8 -*-
"""
araclar_dosya_analiz.py — PDF, Excel, CSV ve Görüntü analiz araçları.

Bagimsiz kullanilabilir; opsiyonel kutuphaneler yoksa graceful degrade yapar:
  - PDF: pdfplumber (tercih) veya PyPDF2
  - Excel: openpyxl veya xlrd
  - CSV: stdlib csv (her zaman mevcut)
  - Goruntu (LLaVA): ollama kutuphanesi veya HTTP API

"""

import csv
import io
import os

# ── Opsiyonel kutuphaneler ─────────────────────────────────────────────────────

try:
    import pdfplumber
    _PDF_VAR = "pdfplumber"
except ImportError:
    pdfplumber = None
    try:
        import PyPDF2
        _PDF_VAR = "PyPDF2"
    except ImportError:
        PyPDF2 = None
        _PDF_VAR = None

try:
    import openpyxl
    _EXCEL_VAR = "openpyxl"
except ImportError:
    openpyxl = None
    try:
        import xlrd
        _EXCEL_VAR = "xlrd"
    except ImportError:
        xlrd = None
        _EXCEL_VAR = None

try:
    import ollama as _ollama_lib
    _OLLAMA_LIB_VAR = True
except ImportError:
    _ollama_lib = None
    _OLLAMA_LIB_VAR = False

# ── Sabitler ──────────────────────────────────────────────────────────────────

OLLAMA_BASE   = "http://localhost:11434"
GORUNTU_MODEL = "llava"
MAKS_SATIR    = 200   # CSV/Excel'de en fazla bu kadar satir goster
MAKS_KARAKTER = 8000  # PDF'te en fazla bu kadar karakter goster


# ── PDF ───────────────────────────────────────────────────────────────────────

def pdf_oku(dosya_yolu: str) -> str:
    """PDF dosyasinin metin icerigini cikartir.

    Args:
        dosya_yolu: PDF dosyasinin tam veya goreceli yolu.

    Returns:
        Metin icerigi veya hata mesaji.
    """
    if not os.path.exists(dosya_yolu):
        return f"[PDF Hatası]: '{dosya_yolu}' bulunamadı."

    if not dosya_yolu.lower().endswith(".pdf"):
        return "[PDF Hatası]: Dosya .pdf uzantılı değil."

    if _PDF_VAR == "pdfplumber":
        try:
            with pdfplumber.open(dosya_yolu) as pdf:
                sayfalar = []
                for i, sayfa in enumerate(pdf.pages, 1):
                    metin = sayfa.extract_text() or ""
                    if metin.strip():
                        sayfalar.append(f"[Sayfa {i}]\n{metin.strip()}")
                tam_metin = "\n\n".join(sayfalar)
                if not tam_metin.strip():
                    return "[PDF]: Metin çıkarılamadı (taranmış/görsel PDF olabilir)."
                if len(tam_metin) > MAKS_KARAKTER:
                    tam_metin = tam_metin[:MAKS_KARAKTER] + f"\n...[{len(tam_metin)} karakter, ilk {MAKS_KARAKTER} gösterildi]"
                return f"[PDF — {len(pdf.pages)} sayfa]\n{tam_metin}"
        except Exception as e:
            return f"[PDF Hatası]: pdfplumber: {e}"

    if _PDF_VAR == "PyPDF2":
        try:
            with open(dosya_yolu, "rb") as f:
                okuyucu = PyPDF2.PdfReader(f)
                sayfalar = []
                for i, sayfa in enumerate(okuyucu.pages, 1):
                    metin = sayfa.extract_text() or ""
                    if metin.strip():
                        sayfalar.append(f"[Sayfa {i}]\n{metin.strip()}")
                tam_metin = "\n\n".join(sayfalar)
                if not tam_metin.strip():
                    return "[PDF]: Metin çıkarılamadı."
                if len(tam_metin) > MAKS_KARAKTER:
                    tam_metin = tam_metin[:MAKS_KARAKTER] + f"\n...[kırpıldı]"
                return f"[PDF — {len(okuyucu.pages)} sayfa]\n{tam_metin}"
        except Exception as e:
            return f"[PDF Hatası]: PyPDF2: {e}"

    return "[PDF]: pdfplumber veya PyPDF2 kurulu değil. 'pip install pdfplumber' çalıştırın."


# ── Excel ─────────────────────────────────────────────────────────────────────

def excel_oku(dosya_yolu: str, sayfa: str = "") -> str:
    """Excel (.xlsx / .xls) dosyasini okur.

    Args:
        dosya_yolu: Excel dosyasinin yolu.
        sayfa:      Sayfa adi veya indeksi. Bos birakılirsa ilk sayfa alinir.

    Returns:
        CSV benzeri metin veya hata mesaji.
    """
    if not os.path.exists(dosya_yolu):
        return f"[Excel Hatası]: '{dosya_yolu}' bulunamadı."

    uzanti = os.path.splitext(dosya_yolu)[1].lower()
    if uzanti not in {".xlsx", ".xls", ".xlsm"}:
        return "[Excel Hatası]: .xlsx/.xls/.xlsm uzantısı gerekli."

    if _EXCEL_VAR == "openpyxl":
        try:
            wb = openpyxl.load_workbook(dosya_yolu, read_only=True, data_only=True)
            if sayfa:
                ws = wb[sayfa] if sayfa in wb.sheetnames else wb.active
            else:
                ws = wb.active
            satirlar = []
            for satir in ws.iter_rows(values_only=True):
                hucre_metinleri = [str(h) if h is not None else "" for h in satir]
                satirlar.append("\t".join(hucre_metinleri))
            wb.close()
            if not satirlar:
                return "[Excel]: Boş sayfa."
            if len(satirlar) > MAKS_SATIR:
                ek = f"\n...[toplam {len(satirlar)} satır, ilk {MAKS_SATIR} gösterildi]"
                satirlar = satirlar[:MAKS_SATIR]
            else:
                ek = ""
            return f"[Excel — '{ws.title}' sayfası, {len(satirlar)} satır]\n" + "\n".join(satirlar) + ek
        except Exception as e:
            return f"[Excel Hatası]: openpyxl: {e}"

    if _EXCEL_VAR == "xlrd":
        try:
            wb = xlrd.open_workbook(dosya_yolu)
            ws = wb.sheet_by_name(sayfa) if sayfa and sayfa in wb.sheet_names() else wb.sheet_by_index(0)
            satirlar = []
            for r in range(ws.nrows):
                hucre_metinleri = [str(ws.cell_value(r, c)) for c in range(ws.ncols)]
                satirlar.append("\t".join(hucre_metinleri))
            if len(satirlar) > MAKS_SATIR:
                ek = f"\n...[toplam {len(satirlar)} satır, ilk {MAKS_SATIR} gösterildi]"
                satirlar = satirlar[:MAKS_SATIR]
            else:
                ek = ""
            return f"[Excel — '{ws.name}' sayfası]\n" + "\n".join(satirlar) + ek
        except Exception as e:
            return f"[Excel Hatası]: xlrd: {e}"

    return "[Excel]: openpyxl veya xlrd kurulu değil. 'pip install openpyxl' çalıştırın."


# ── CSV ───────────────────────────────────────────────────────────────────────

def csv_oku(dosya_yolu: str, ayirici: str = ",") -> str:
    """CSV dosyasini okur ve ozet tablo olarak dondurur.

    Args:
        dosya_yolu: CSV dosyasinin yolu.
        ayirici:    Alan ayirici karakter (varsayilan virgul).

    Returns:
        Tablo metni veya hata mesaji.
    """
    if not os.path.exists(dosya_yolu):
        return f"[CSV Hatası]: '{dosya_yolu}' bulunamadı."

    try:
        # Ayirici tespiti
        if not ayirici or ayirici == "auto":
            with open(dosya_yolu, "r", encoding="utf-8", errors="replace") as f:
                ornek = f.read(4096)
            lehcelen = csv.Sniffer().sniff(ornek, delimiters=",;\t|")
            ayirici = lehcelen.delimiter
    except Exception:
        ayirici = ","

    try:
        satirlar = []
        with open(dosya_yolu, "r", encoding="utf-8", errors="replace") as f:
            okuyucu = csv.reader(f, delimiter=ayirici)
            for i, satir in enumerate(okuyucu):
                if i >= MAKS_SATIR:
                    satirlar.append(f"...[{i} satır okundu, geri kalanlar atlandı]")
                    break
                satirlar.append("\t".join(satir))

        if not satirlar:
            return "[CSV]: Boş dosya."
        return f"[CSV — {len(satirlar)} satır, ayirici='{ayirici}']\n" + "\n".join(satirlar)
    except Exception as e:
        return f"[CSV Hatası]: {e}"


# ── LLaVA Görüntü Analizi ─────────────────────────────────────────────────────

def goruntu_analiz(goruntu_yolu: str, soru: str = "") -> str:
    """Goruntu analiz eder (OpenRouter vision uzerinden)."""
    if not os.path.exists(goruntu_yolu):
        return f"[Görüntü Hatası]: '{goruntu_yolu}' bulunamadı."
    try:
        from reymen.arac.araclar_goruntu import vision_analiz
        return vision_analiz(kaynak=goruntu_yolu, soru=soru or "Bu görselde ne var, detaylı açıkla.")
    except Exception as e:
        return f"[Görüntü Hatası]: {e}"


def _goruntu_http(goruntu_yolu: str, prompt: str, onceki_hata: str) -> str:
    """Artik kullanilmiyor (yerine vision_analiz kullan)."""
    return goruntu_analiz(goruntu_yolu, prompt)


# ── Dosya Tipi Tespiti (tek API noktası) ─────────────────────────────────────

def dosya_analiz(dosya_yolu: str, ek_parametre: str = "") -> str:
    """Dosya uzantisina gore dogru arac cagrisini yapar.

    Args:
        dosya_yolu:    Analiz edilecek dosya.
        ek_parametre:  PDF/Excel icin soru veya sayfa adi; CSV icin ayirici.

    Returns:
        Analiz sonucu.
    """
    if not dosya_yolu:
        return "[Dosya Analiz]: Dosya yolu boş."

    uzanti = os.path.splitext(dosya_yolu)[1].lower()

    if uzanti == ".pdf":
        return pdf_oku(dosya_yolu)
    if uzanti in {".xlsx", ".xls", ".xlsm"}:
        return excel_oku(dosya_yolu, sayfa=ek_parametre)
    if uzanti == ".csv":
        return csv_oku(dosya_yolu, ayirici=ek_parametre or ",")
    if uzanti in {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".gif"}:
        return goruntu_analiz(dosya_yolu, soru=ek_parametre)
    return (
        f"[Dosya Analiz]: '{uzanti}' uzantısı desteklenmiyor.\n"
        f"Desteklenenler: .pdf, .xlsx, .xls, .csv, .jpg, .jpeg, .png, .webp"
    )


if __name__ == "__main__":
    print("=== Dosya Analiz Araçları Testi ===")
    print(f"PDF:   {_PDF_VAR or 'kurulu değil'}")
    print(f"Excel: {_EXCEL_VAR or 'kurulu değil'}")
    print(f"CSV:   stdlib (her zaman mevcut)")
    print(f"LLaVA: {'ollama kütüphanesi' if _OLLAMA_LIB_VAR else 'HTTP API'} üzerinden\n")

    # CSV testi (stdlib ile her zaman çalışır)
    import tempfile
    tmp = tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False, encoding="utf-8")
    tmp.write("ad,yas,sehir\nAhmet,30,Istanbul\nFatma,25,Ankara\nMehmet,35,Izmir\n")
    tmp.close()
    print(csv_oku(tmp.name))
    os.unlink(tmp.name)

    # Görüntü testi — dosya yoksa graceful
    print(goruntu_analiz("olmayan_resim.jpg", "Ne var?"))

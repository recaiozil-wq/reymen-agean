"""
İçerik okuma ve çıkarma aracı.
PDF, DOCX, XLSX, ZIP içinden metin çıkarma.
"""
import os
import logging
logger = logging.getLogger(__name__)


def _turu_bul(dosya_yolu):
    """Dosya türünü bul."""
    _, ext = os.path.splitext(dosya_yolu)
    ext = ext.lower()
    tur_haritasi = {
        '.pdf': 'PDF',
        '.docx': 'DOCX',
        '.xlsx': 'XLSX',
        '.zip': 'ZIP',
        '.txt': 'TXT',
        '.md': 'MD',
        '.csv': 'CSV',
        '.json': 'JSON',
    }
    return tur_haritasi.get(ext, f"Bilinmeyen ({ext})")


def _pdf_oku(dosya_yolu):
    """PDF'den metin çıkar."""
    try:
        import PyPDF2
        metin = []
        with open(dosya_yolu, 'rb') as f:
            okuyucu = PyPDF2.PdfReader(f)
            for sayfa in okuyucu.pages:
                metin.append(sayfa.extract_text())
        return "\n".join(metin), None
    except ImportError:
        try:
            import pdfminer
            from pdfminer.high_level import extract_text
            return extract_text(dosya_yolu), None
        except ImportError:
            return None, "PyPDF2 veya pdfminer modülü gerekli."


def _docx_oku(dosya_yolu):
    """DOCX'den metin çıkar."""
    try:
        from docx import Document
        doc = Document(dosya_yolu)
        metin = [p.text for p in doc.paragraphs]
        return "\n".join(metin), None
    except ImportError:
        return None, "python-docx modülü gerekli."


def _xlsx_oku(dosya_yolu):
    """XLSX'den metin çıkar."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(dosya_yolu, read_only=True, data_only=True)
        metin = []
        for sayfa in wb.sheetnames:
            ws = wb[sayfa]
            metin.append(f"=== Sayfa: {sayfa} ===")
            for row in ws.iter_rows(values_only=True):
                satir = " | ".join([str(c) if c is not None else "" for c in row])
                if satir.strip():
                    metin.append(satir)
        return "\n".join(metin), None
    except ImportError:
        return None, "openpyxl modülü gerekli."


def _zip_oku(dosya_yolu):
    """ZIP içindeki dosyaları listele ve metin dosyalarını oku."""
    import zipfile
    metin = []
    with zipfile.ZipFile(dosya_yolu, 'r') as z:
        for info in z.infolist():
            if not info.is_dir():
                metin.append(f"[{info.filename}] ({info.file_size} bayt)")
                try:
                    with z.open(info) as f:
                        icerik = f.read().decode('utf-8', errors='replace')[:500]
                        metin.append(icerik)
                except Exception:
                    metin.append("(binary/okunamaz)")
    return "\n".join(metin), None


def run(islem='turu_bul', dosya_yolu=None, **kwargs):
    """
    İçerik okuma ve çıkarma.

    Parametreler:
        islem (str): 'oku', 'cikar' veya 'turu_bul'
        dosya_yolu (str): Dosya yolu

    Returns:
        str: İşlem sonucu.
    """
    try:
        if not dosya_yolu:
            return "Hata: 'dosya_yolu' parametresi zorunludur."

        if not os.path.exists(dosya_yolu):
            return f"Hata: Dosya bulunamadı: {dosya_yolu}"

        if islem == 'turu_bul':
            return f"Dosya türü: {_turu_bul(dosya_yolu)}"

        elif islem in ('oku', 'cikar'):
            tur = _turu_bul(dosya_yolu)
            if tur == 'PDF':
                icerik, hata = _pdf_oku(dosya_yolu)
            elif tur == 'DOCX':
                icerik, hata = _docx_oku(dosya_yolu)
            elif tur == 'XLSX':
                icerik, hata = _xlsx_oku(dosya_yolu)
            elif tur == 'ZIP':
                icerik, hata = _zip_oku(dosya_yolu)
            else:
                try:
                    with open(dosya_yolu, 'r', encoding='utf-8', errors='replace') as f:
                        icerik = f.read()
                    hata = None
                except Exception as e:
                    icerik, hata = None, str(e)

            if hata:
                return f"Okuma hatası ({tur}): {hata}"
            return f"=== {dosya_yolu} ({tur}) ===\n{icerik[:5000]}" + ("\n...(devamı kesildi)" if len(icerik) > 5000 else "")

        else:
            return f"Hata: Geçersiz işlem '{islem}'. 'oku', 'cikar' veya 'turu_bul' kullanın."

    except Exception as e:
        return f"Okuma/çıkarma hatası: {str(e)}"

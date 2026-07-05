# -*- coding: utf-8 -*-
"""
araclar_dosya_analiz.py â€” PDF, Excel, CSV ve Görüntü analiz araçlarÄ±.

Bagimsiz kullanilabilir; opsiyonel kutuphaneler yoksa graceful degrade yapar:
  - PDF: pdfplumber (tercih) veya PyPDF2
  - Excel: openpyxl veya xlrd
  - CSV: stdlib csv (her zaman mevcut)
  - Goruntu (LLaVA): ollama kutuphanesi veya HTTP API

"""

import csv
import io
import os

# â”€â”€ Opsiyonel kutuphaneler â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

try:
    import pdfplumber

    _PDF_VAR = "pdfplumber"
except ImportError:
    pdfplumber = None
    try:
        import PyPDF2

        _PDF_VAR = "PyPDF2"
    except ImportError:
        PyPDF2 = None
        _PDF_VAR = None

try:
    import openpyxl

    _EXCEL_VAR = "openpyxl"
except ImportError:
    openpyxl = None
    try:
        import xlrd

        _EXCEL_VAR = "xlrd"
    except ImportError:
        xlrd = None
        _EXCEL_VAR = None

try:
    import ollama as _ollama_lib

    _OLLAMA_LIB_VAR = True
except ImportError:
    _ollama_lib = None
    _OLLAMA_LIB_VAR = False

# â”€â”€ Sabitler â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

OLLAMA_BASE = "http://localhost:11434"
GORUNTU_MODEL = "llava"
MAKS_SATIR = 200  # CSV/Excel'de en fazla bu kadar satir goster
MAKS_KARAKTER = 8000  # PDF'te en fazla bu kadar karakter goster


# â”€â”€ PDF â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€


def pdf_oku(dosya_yolu: str) -> str:
    """PDF dosyasinin metin icerigini cikartir.

    Args:
        dosya_yolu: PDF dosyasinin tam veya goreceli yolu.

    Returns:
        Metin icerigi veya hata mesaji.
    """
    if not os.path.exists(dosya_yolu):
        return f"[PDF HatasÄ±]: '{dosya_yolu}' bulunamadÄ±."

    if not dosya_yolu.lower().endswith(".pdf"):
        return "[PDF HatasÄ±]: Dosya .pdf uzantÄ±lÄ± deÄŸil."

    if _PDF_VAR == "pdfplumber":
        try:
            with pdfplumber.open(dosya_yolu) as pdf:
                sayfalar = []
                for i, sayfa in enumerate(pdf.pages, 1):
                    metin = sayfa.extract_text() or ""
                    if metin.strip():
                        sayfalar.append(f"[Sayfa {i}]\n{metin.strip()}")
                tam_metin = "\n\n".join(sayfalar)
                if not tam_metin.strip():
                    return "[PDF]: Metin çÄ±karÄ±lamadÄ± (taranmÄ±ÅŸ/görsel PDF olabilir)."
                if len(tam_metin) > MAKS_KARAKTER:
                    tam_metin = (
                        tam_metin[:MAKS_KARAKTER]
                        + f"\n...[{len(tam_metin)} karakter, ilk {MAKS_KARAKTER} gösterildi]"
                    )
                return f"[PDF â€” {len(pdf.pages)} sayfa]\n{tam_metin}"
        except Exception as e:
            return f"[PDF HatasÄ±]: pdfplumber: {e}"

    if _PDF_VAR == "PyPDF2":
        try:
            with open(dosya_yolu, "rb") as f:
                okuyucu = PyPDF2.PdfReader(f)
                sayfalar = []
                for i, sayfa in enumerate(okuyucu.pages, 1):
                    metin = sayfa.extract_text() or ""
                    if metin.strip():
                        sayfalar.append(f"[Sayfa {i}]\n{metin.strip()}")
                tam_metin = "\n\n".join(sayfalar)
                if not tam_metin.strip():
                    return "[PDF]: Metin çÄ±karÄ±lamadÄ±."
                if len(tam_metin) > MAKS_KARAKTER:
                    tam_metin = tam_metin[:MAKS_KARAKTER] + f"\n...[kÄ±rpÄ±ldÄ±]"
                return f"[PDF â€” {len(okuyucu.pages)} sayfa]\n{tam_metin}"
        except Exception as e:
            return f"[PDF HatasÄ±]: PyPDF2: {e}"

    return "[PDF]: pdfplumber veya PyPDF2 kurulu deÄŸil. 'pip install pdfplumber' çalÄ±ÅŸtÄ±rÄ±n."


# â”€â”€ Excel â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€


def excel_oku(dosya_yolu: str, sayfa: str = "") -> str:
    """Excel (.xlsx / .xls) dosyasini okur.

    Args:
        dosya_yolu: Excel dosyasinin yolu.
        sayfa:      Sayfa adi veya indeksi. Bos birakÄ±lirsa ilk sayfa alinir.

    Returns:
        CSV benzeri metin veya hata mesaji.
    """
    if not os.path.exists(dosya_yolu):
        return f"[Excel HatasÄ±]: '{dosya_yolu}' bulunamadÄ±."

    uzanti = os.path.splitext(dosya_yolu)[1].lower()
    if uzanti not in {".xlsx", ".xls", ".xlsm"}:
        return "[Excel HatasÄ±]: .xlsx/.xls/.xlsm uzantÄ±sÄ± gerekli."

    if _EXCEL_VAR == "openpyxl":
        try:
            wb = openpyxl.load_workbook(dosya_yolu, read_only=True, data_only=True)
            if sayfa:
                ws = wb[sayfa] if sayfa in wb.sheetnames else wb.active
            else:
                ws = wb.active
            satirlar = []
            for satir in ws.iter_rows(values_only=True):
                hucre_metinleri = [str(h) if h is not None else "" for h in satir]
                satirlar.append("\t".join(hucre_metinleri))
            wb.close()
            if not satirlar:
                return "[Excel]: BoÅŸ sayfa."
            if len(satirlar) > MAKS_SATIR:
                ek = f"\n...[toplam {len(satirlar)} satÄ±r, ilk {MAKS_SATIR} gösterildi]"
                satirlar = satirlar[:MAKS_SATIR]
            else:
                ek = ""
            return (
                f"[Excel â€” '{ws.title}' sayfasÄ±, {len(satirlar)} satÄ±r]\n"
                + "\n".join(satirlar)
                + ek
            )
        except Exception as e:
            return f"[Excel HatasÄ±]: openpyxl: {e}"

    if _EXCEL_VAR == "xlrd":
        try:
            wb = xlrd.open_workbook(dosya_yolu)
            ws = (
                wb.sheet_by_name(sayfa)
                if sayfa and sayfa in wb.sheet_names()
                else wb.sheet_by_index(0)
            )
            satirlar = []
            for r in range(ws.nrows):
                hucre_metinleri = [str(ws.cell_value(r, c)) for c in range(ws.ncols)]
                satirlar.append("\t".join(hucre_metinleri))
            if len(satirlar) > MAKS_SATIR:
                ek = f"\n...[toplam {len(satirlar)} satÄ±r, ilk {MAKS_SATIR} gösterildi]"
                satirlar = satirlar[:MAKS_SATIR]
            else:
                ek = ""
            return f"[Excel â€” '{ws.name}' sayfasÄ±]\n" + "\n".join(satirlar) + ek
        except Exception as e:
            return f"[Excel HatasÄ±]: xlrd: {e}"

    return (
        "[Excel]: openpyxl veya xlrd kurulu deÄŸil. 'pip install openpyxl' çalÄ±ÅŸtÄ±rÄ±n."
    )


# â”€â”€ CSV â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€


def csv_oku(dosya_yolu: str, ayirici: str = ",") -> str:
    """CSV dosyasini okur ve ozet tablo olarak dondurur.

    Args:
        dosya_yolu: CSV dosyasinin yolu.
        ayirici:    Alan ayirici karakter (varsayilan virgul).

    Returns:
        Tablo metni veya hata mesaji.
    """
    if not os.path.exists(dosya_yolu):
        return f"[CSV HatasÄ±]: '{dosya_yolu}' bulunamadÄ±."

    try:
        # Ayirici tespiti
        if not ayirici or ayirici == "auto":
            with open(dosya_yolu, "r", encoding="utf-8", errors="replace") as f:
                ornek = f.read(4096)
            lehcelen = csv.Sniffer().sniff(ornek, delimiters=",;\t|")
            ayirici = lehcelen.delimiter
    except Exception:
        ayirici = ","

    try:
        satirlar = []
        with open(dosya_yolu, "r", encoding="utf-8", errors="replace") as f:
            okuyucu = csv.reader(f, delimiter=ayirici)
            for i, satir in enumerate(okuyucu):
                if i >= MAKS_SATIR:
                    satirlar.append(f"...[{i} satÄ±r okundu, geri kalanlar atlandÄ±]")
                    break
                satirlar.append("\t".join(satir))

        if not satirlar:
            return "[CSV]: BoÅŸ dosya."
        return f"[CSV â€” {len(satirlar)} satÄ±r, ayirici='{ayirici}']\n" + "\n".join(
            satirlar
        )
    except Exception as e:
        return f"[CSV HatasÄ±]: {e}"


# â”€â”€ LLaVA Görüntü Analizi â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€


def goruntu_analiz(goruntu_yolu: str, soru: str = "") -> str:
    """Goruntu analiz eder (OpenRouter vision uzerinden)."""
    if not os.path.exists(goruntu_yolu):
        return f"[Görüntü HatasÄ±]: '{goruntu_yolu}' bulunamadÄ±."
    try:
        from reymen.arac.araclar_goruntu import vision_analiz

        return vision_analiz(
            kaynak=goruntu_yolu, soru=soru or "Bu görselde ne var, detaylÄ± açÄ±kla."
        )
    except Exception as e:
        return f"[Görüntü HatasÄ±]: {e}"


def _goruntu_http(goruntu_yolu: str, prompt: str, onceki_hata: str) -> str:
    """Artik kullanilmiyor (yerine vision_analiz kullan)."""
    return goruntu_analiz(goruntu_yolu, prompt)


# â”€â”€ Dosya Tipi Tespiti (tek API noktasÄ±) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€


def dosya_analiz(dosya_yolu: str, ek_parametre: str = "") -> str:
    """Dosya uzantisina gore dogru arac cagrisini yapar.

    Args:
        dosya_yolu:    Analiz edilecek dosya.
        ek_parametre:  PDF/Excel icin soru veya sayfa adi; CSV icin ayirici.

    Returns:
        Analiz sonucu.
    """
    if not dosya_yolu:
        return "[Dosya Analiz]: Dosya yolu boÅŸ."

    uzanti = os.path.splitext(dosya_yolu)[1].lower()

    if uzanti == ".pdf":
        return pdf_oku(dosya_yolu)
    if uzanti in {".xlsx", ".xls", ".xlsm"}:
        return excel_oku(dosya_yolu, sayfa=ek_parametre)
    if uzanti == ".csv":
        return csv_oku(dosya_yolu, ayirici=ek_parametre or ",")
    if uzanti in {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".gif"}:
        return goruntu_analiz(dosya_yolu, soru=ek_parametre)
    return (
        f"[Dosya Analiz]: '{uzanti}' uzantÄ±sÄ± desteklenmiyor.\n"
        f"Desteklenenler: .pdf, .xlsx, .xls, .csv, .jpg, .jpeg, .png, .webp"
    )


if __name__ == "__main__":
    print("=== Dosya Analiz AraçlarÄ± Testi ===")
    print(f"PDF:   {_PDF_VAR or 'kurulu deÄŸil'}")
    print(f"Excel: {_EXCEL_VAR or 'kurulu deÄŸil'}")
    print(f"CSV:   stdlib (her zaman mevcut)")
    print(
        f"LLaVA: {'ollama kütüphanesi' if _OLLAMA_LIB_VAR else 'HTTP API'} üzerinden\n"
    )

    # CSV testi (stdlib ile her zaman çalÄ±ÅŸÄ±r)
    import tempfile

    tmp = tempfile.NamedTemporaryFile(
        mode="w", suffix=".csv", delete=False, encoding="utf-8"
    )
    tmp.write("ad,yas,sehir\nAhmet,30,Istanbul\nFatma,25,Ankara\nMehmet,35,Izmir\n")
    tmp.close()
    print(csv_oku(tmp.name))
    os.unlink(tmp.name)

    # Görüntü testi â€” dosya yoksa graceful
    print(goruntu_analiz("olmayan_resim.jpg", "Ne var?"))

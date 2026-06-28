from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

# Sheet 1: Yatırım Şablonu
ws = wb.active
ws.title = "Yatırım Şablonu"

# Styles
dark_blue = "1F4E79"
turquoise = "00BCD4"
zebra1 = "E8F6F8"
zebra2 = "FFFFFF"
header_fill = PatternFill(start_color=dark_blue, end_color=dark_blue, fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Title
ws.merge_cells("A1:I1")
ws["A1"] = "Kentsel Donusum Yatirim Analiz Sablonu"
ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws["A1"].fill = PatternFill(start_color="0D3B66", end_color="0D3B66", fill_type="solid")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# Headers
headers = [
    "Yatirim Secimi",
    "Toplam Yatirim (TL)",
    "Gayrimenkul Deger Artisi (%)",
    "Yatirimin Amaci",
    "Risk Tercihi",
    "Finansman",
    "Sure (Ay)",
]
for col in range(1, 8):
    cell = ws.cell(row=3, column=col, value=headers[col - 1])
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Investment Choices
rows = [
    ["Konut", "2,000,000 TL", "%12", "Kiralik Konut", "Orta", "Kendi Mevc", "36"],
    ["Ticari", "3,000,000 TL", "%16", "Acik Ofis/Kafe/Restorant", "Yuksek", "Kredi", "48"],
    ["Turistik", "4,000,000 TL", "%18", "Butik Otel/Otel", "Orta-Yüksek", "Ortak", "60"],
    ["Kentsel Donusum", "1,500,000 TL", "%22", "Arsa/Bina", "Düşük-Orta", "Nakit", "24"],
    ["Lojistik", "2,500,000 TL", "%8", "Depo/Ist", "Düşük", "Kendi Mevc", "42"],
]

for row_data in rows:
    row_num = ws.max_row + 1
    fill_color = zebra1 if row_num % 2 == 0 else zebra2
    row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    for col, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_num, column=col)
        cell.value = value
        cell.border = thin_border
        cell.fill = row_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

# Footer
ws.merge_cells("A8:G8")
ws["A8"] = "— Yatirima gore secenekleri degerlendip dogru karari verebilmeniz için —"
ws["A8"].font = Font(color="888888", italic=True)
ws["A8"].alignment = Alignment(horizontal="center")

# Column widths (A-G)
ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 28
ws.column_dimensions["E"].width = 16
ws.column_dimensions["F"].width = 20
ws.column_dimensions["G"].width = 12

output_path = "Yatirim_Secim_Sablonu.xlsx"
wb.save(output_path)
print("OK", output_path)
